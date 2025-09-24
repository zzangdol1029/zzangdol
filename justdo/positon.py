import requests
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import numpy as np
import os

# 파일 경로
INPUT_FILE = "C:/zzangdol/justdo/250924_홈커밍데이_자리배치도 (1).xlsx"
OUTPUT_FILE = "C:/zzangdol/justdo/250924_홈커밍데이_자리배치도_with_position.xlsx"

# API URL
BASE_URL = "https://sd.solideos.com/api/org/user/sort/list"

# 인증 정보 (실제 쿠키 값)
HEADERS = {
    'Content-Type': 'application/json',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

# 실제 쿠키 정보
COOKIES = {
    'GOSSOcookie': '82e2d0c2-d230-45bc-a308-ce676d54f147',
    'IsCookieActived': 'true',
    'OcxUpload': 'off',
    'PRW': 'show',
    'PUW': 'show',
    'fileDownload': 'true',
    'isAttachAreaFolded': 'false',
    'isDoCareLoungeCloseClick': 'true',
    'reflash_time': '-1'
}

def create_team_charts(team_stats, all_positions):
    """조별 직급 통계 차트들을 생성"""
    # 한글 폰트 설정 (한글이 깨지지 않도록)
    plt.rcParams['font.family'] = 'Malgun Gothic'  # Windows
    plt.rcParams['axes.unicode_minus'] = False
    
    # 차트 저장 폴더 생성
    chart_dir = "C:/zzangdol/justdo/charts"
    if not os.path.exists(chart_dir):
        os.makedirs(chart_dir)
    
    chart_files = []
    
    # 1. 전체 조별 비교 막대 그래프
    plt.figure(figsize=(15, 8))
    
    # 각 직급별로 조들의 데이터 준비
    teams = sorted(team_stats.keys())
    x = np.arange(len(teams))
    width = 0.8 / len(all_positions)
    
    colors = plt.cm.Set3(np.linspace(0, 1, len(all_positions)))
    
    for i, position in enumerate(all_positions):
        counts = [team_stats[team].get(position, 0) for team in teams]
        plt.bar(x + i * width, counts, width, label=position, color=colors[i])
    
    plt.xlabel('조')
    plt.ylabel('인원수')
    plt.title('조별 직급 분포')
    plt.xticks(x + width * (len(all_positions) - 1) / 2, [f'{team}조' for team in teams])
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    
    chart_file = f"{chart_dir}/team_comparison.png"
    plt.savefig(chart_file, dpi=300, bbox_inches='tight')
    chart_files.append(chart_file)
    plt.close()
    
    # 2. 전체 직급 분포 파이 차트
    plt.figure(figsize=(10, 8))
    
    total_by_position = {}
    for team_data in team_stats.values():
        for position, count in team_data.items():
            total_by_position[position] = total_by_position.get(position, 0) + count
    
    positions = list(total_by_position.keys())
    counts = list(total_by_position.values())
    
    plt.pie(counts, labels=positions, autopct='%1.1f%%', startangle=90)
    plt.title('전체 직급 분포')
    plt.axis('equal')
    plt.tight_layout()
    
    chart_file = f"{chart_dir}/total_pie.png"
    plt.savefig(chart_file, dpi=300, bbox_inches='tight')
    chart_files.append(chart_file)
    plt.close()
    
    return chart_files

def insert_charts_to_excel(wb, chart_files):
    """차트들을 엑셀에 삽입"""
    # 차트 시트 생성
    ws_charts = wb.create_sheet("직급통계_차트")
    
    current_row = 1
    
    for i, chart_file in enumerate(chart_files):
        try:
            # 이미지 삽입
            img = Image(chart_file)
            
            # 이미지 크기 조정
            img.width = 600
            img.height = 400
            
            # 위치 설정
            img.anchor = f'A{current_row}'
            current_row += 25  # 다음 차트를 위한 공간
            
            ws_charts.add_image(img)
            
        except Exception as e:
            print(f"차트 삽입 오류: {chart_file} - {e}")
    
    return ws_charts

def get_position(name: str) -> str:
    """주어진 이름으로 API 요청 후 position 반환"""
    if pd.isna(name) or name.strip() == "":
        return ""
    
    # 테스트 모드 비활성화 - 실제 API 사용
    TEST_MODE = False
    if TEST_MODE:
        import random
        positions = ["사원", "주임", "대리", "과장", "차장", "부장", "이사", "상무", "전무", "대표"]
        position = random.choice(positions)
        print(f"{name}: {position} (테스트 데이터)")
        return position
    
    params = {
        "keyword": name.strip(),
        "page": 0,
        "offset": 30,
        "nodeType": "org"
    }
    try:
        # 헤더와 쿠키를 포함한 요청
        response = requests.get(
            BASE_URL, 
            params=params, 
            headers=HEADERS,
            cookies=COOKIES,  # 실제 쿠키 사용
            timeout=10
        )
        
        if response.status_code == 401:
            print(f"{name}: 인증 오류 (401) - 쿠키가 만료되었을 수 있습니다")
            return "AUTH_ERROR"
        
        response.raise_for_status()
        data = response.json()

        # 실제 API 구조에 따라 수정 필요
        if data and "data" in data and len(data["data"]) > 0:
            position = data["data"][0].get("position", "")
            print(f"{name}: {position}")
            return position
        else:
            print(f"{name}: N/A")
            return "N/A"
    except requests.exceptions.HTTPError as e:
        if "401" in str(e):
            print(f"{name}: 인증 실패 - 쿠키가 만료되었을 수 있습니다")
            return "AUTH_REQUIRED"
        else:
            print(f"{name}: HTTP 오류 - {e}")
            return "HTTP_ERROR"
    except Exception as e:
        print(f"Error fetching {name}: {e}")
        return "ERROR"

# 엑셀 불러오기
print("엑셀 파일 읽는 중...")
df = pd.read_excel(INPUT_FILE)

print("데이터 구조:")
print(f"행 수: {len(df)}, 열 수: {len(df.columns)}")
print(f"컬럼명: {df.columns.tolist()}")

# 새로운 워크북 생성
wb = Workbook()

# 원본 데이터 시트
ws_original = wb.active
ws_original.title = "원본_자리배치"

# 직급 정보 시트
ws_position = wb.create_sheet("직급_정보")

# 조별 직급 통계 시트
ws_stats = wb.create_sheet("조별_직급통계")

print("\n원본 데이터를 새 워크북에 복사 중...")
# 원본 데이터 복사
for i, row in df.iterrows():
    for j, value in enumerate(row):
        ws_original.cell(row=i+1, column=j+1, value=value)

print("\n직급 정보 조회 중...")
# 조별 직급 통계를 위한 딕셔너리
team_stats = {}

# 직급 정보 조회 및 저장
for i, row in df.iterrows():
    team_num = i + 1  # 조 번호 (1부터 시작)
    team_stats[team_num] = {}
    
    for j, name in enumerate(row):
        if pd.notna(name) and str(name).strip() != "":
            position = get_position(str(name))
            # 원본 이름은 그대로 저장
            ws_position.cell(row=i+1, column=j*2+1, value=str(name))
            # 바로 옆 열에 직급만 저장
            ws_position.cell(row=i+1, column=j*2+2, value=position)
            
            # 조별 통계에 추가
            if position and position not in ["N/A", "ERROR", "AUTH_ERROR", "HTTP_ERROR", "AUTH_REQUIRED"]:
                if position in team_stats[team_num]:
                    team_stats[team_num][position] += 1
                else:
                    team_stats[team_num][position] = 1
        else:
            # 빈 셀인 경우 두 열 모두 비워둠
            ws_position.cell(row=i+1, column=j*2+1, value="")
            ws_position.cell(row=i+1, column=j*2+2, value="")

print("\n조별 직급 통계 생성 중...")
# 조별 직급 통계 시트 작성
# 헤더 작성
ws_stats.cell(row=1, column=1, value="조")
all_positions = set()
for team_data in team_stats.values():
    all_positions.update(team_data.keys())

all_positions = sorted(list(all_positions))

# 직급별 헤더 추가
for idx, position in enumerate(all_positions):
    ws_stats.cell(row=1, column=idx+2, value=position)

ws_stats.cell(row=1, column=len(all_positions)+2, value="총원")

# 각 조별 데이터 입력
for team_num in sorted(team_stats.keys()):
    ws_stats.cell(row=team_num+1, column=1, value=f"{team_num}조")
    
    total_count = 0
    for idx, position in enumerate(all_positions):
        count = team_stats[team_num].get(position, 0)
        ws_stats.cell(row=team_num+1, column=idx+2, value=count)
        total_count += count
    
    ws_stats.cell(row=team_num+1, column=len(all_positions)+2, value=total_count)

# 전체 합계 행 추가
ws_stats.cell(row=len(team_stats)+2, column=1, value="전체")
for idx, position in enumerate(all_positions):
    total = sum(team_data.get(position, 0) for team_data in team_stats.values())
    ws_stats.cell(row=len(team_stats)+2, column=idx+2, value=total)

# 전체 총원
grand_total = sum(sum(team_data.values()) for team_data in team_stats.values())
ws_stats.cell(row=len(team_stats)+2, column=len(all_positions)+2, value=grand_total)

# 차트 생성 및 엑셀에 삽입
chart_files = create_team_charts(team_stats, all_positions)
ws_charts = insert_charts_to_excel(wb, chart_files)

# 엑셀 저장
print(f"\n결과를 저장 중: {OUTPUT_FILE}")
wb.save(OUTPUT_FILE)

print(f"완료! 엑셀 파일 저장: {OUTPUT_FILE}")
print("- '원본_자리배치' 시트: 원본 데이터")
print("- '직급_정보' 시트: 이름과 직급이 나란히 표시된 자리배치표")
print("- '조별_직급통계' 시트: 각 조별 직급 인원수 및 전체 통계")
print("- '직급통계_차트' 시트: 조별 직급 분포 그래프 및 전체 직급 분포 차트")